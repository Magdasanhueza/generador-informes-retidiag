#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de Informes Retinográficos PDF
Retidiag - 2026

Este programa lee un archivo Excel con datos de pacientes y genera
informes PDF individuales según el diagnóstico.
"""

import os
import sys
import shutil
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY

# Configuración de rutas
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGOS_DIR = os.path.join(BASE_DIR, "imagenes", "logos")
FIRMAS_DIR = os.path.join(BASE_DIR, "imagenes", "firmas")
OUTPUT_DIR = "/Users/magda/Documents/informes retidiag"
PLANTILLA_RESUMEN = os.path.join(BASE_DIR, "plantilla_resumen_pacientes.xls")

# Crear directorio de salida si no existe
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Mapeo de comunas a logos de establecimientos
LOGOS_ESTABLECIMIENTO = {
    "PEÑALOLÉN": "logo_penalolen.jpg",
    "PENALOLEN": "logo_penalolen.jpg",
    "LAS CONDES": "logo_las_condes.png",
    "EL MONTE": "logo_el_monte.png",
    "PROVIDENCIA": "logo_providencia.jpg",
}

# Mapeo de oftalmólogos a firmas
FIRMAS_OFTALMOLOGO = {
    "DR. CONTRERAS": "firma_felipe_contreras.jpg",
    "DRA. ELTIT": "firma_yasmine_eltit.png",
    "YASMINE ELTIT": "firma_yasmine_eltit.png",
}

# Mapeo de TMOs a firmas
FIRMAS_TMO = {
    "FELIPE ROJAS": "firma_tmo_felipe_rojas.jpg",
    "MAURICIO PEREZ": "firma_mauricio_perez.png",
    "MAURICIO PÉREZ": "firma_mauricio_perez.png",
    "JOSEFINA HERRERA": "firma_josefina_herrera.png",
    "JAVIERA COMPAN": "firma_javiera_compan.png",
    "HECTOR VERA": "firma_hector_vera.png",
    "HÉCTOR VERA": "firma_hector_vera.png",
}

# Textos de diagnóstico según resultado
TEXTOS_DIAGNOSTICO = {
    "NORMAL": [
        "- No se observan signos de retinopatía diabética.",
        "- No se observan otras alteraciones retinianas significativas.",
    ],
    "DG NORMAL": [
        "- No se observan signos de retinopatía diabética.",
        "- No se observan otras alteraciones retinianas significativas.",
        "",
        "Este examen fue evaluado por oftalmólogo.",
    ],
    "CATARATA": [
        "- No se logró observar la retina en forma nítida en las múltiples fotografías obtenidas, opacidad de medios, sospecha de cataratas.",
    ],
    "RD": [
        "Se observan signos de retinopatía diabética.",
    ],
    "OTROS": [
        "- No se observan signos de retinopatía diabética, pero se aprecian otras alteraciones:",
    ],
}

# Sugerencias según resultado
SUGERENCIAS = {
    "NORMAL": [
        "- Mantener control metabólico.",
        "- Control en un año.",
    ],
    "DG NORMAL": [
        "- Mantener control metabólico.",
        "- Control en un año.",
    ],
    "CATARATA": [
        "- Se sugiere derivar a oftalmología para evaluación de cataratas.",
    ],
    "RD": [
        "- Se sugiere derivar a oftalmología para evaluación y tratamiento.",
        "- Mantener estricto control metabólico.",
    ],
    "OTROS": [
        "- Se sugiere derivar a oftalmología para evaluación.",
    ],
}

# Derivaciones según resultado
DERIVACIONES = {
    "NORMAL": "FONDO DE OJO ANUAL",
    "DG NORMAL": "FONDO DE OJO ANUAL",
    "CATARATA": "DERIVAR OFTALMOLOGÍA",
    "RD": "DERIVAR OFTALMOLOGÍA",
    "OTROS": "DERIVAR OFTALMOLOGÍA",
}


def crear_estilos():
    """Crea y retorna los estilos para el PDF."""
    styles = getSampleStyleSheet()

    # Estilo para el encabezado de la empresa
    styles.add(ParagraphStyle(
        name='Empresa',
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#333333'),
        alignment=TA_LEFT,
        spaceAfter=1*mm,
    ))

    # Estilo para el título
    styles.add(ParagraphStyle(
        name='Titulo',
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#2c5282'),
        alignment=TA_CENTER,
        spaceBefore=5*mm,
        spaceAfter=8*mm,
    ))

    # Estilo para etiquetas
    styles.add(ParagraphStyle(
        name='Etiqueta',
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.black,
    ))

    # Estilo para valores
    styles.add(ParagraphStyle(
        name='Valor',
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.black,
    ))

    # Estilo para texto del cuerpo
    styles.add(ParagraphStyle(
        name='Cuerpo',
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.black,
        alignment=TA_JUSTIFY,
        spaceBefore=2*mm,
        spaceAfter=2*mm,
        leading=14,
    ))

    # Estilo para subtítulos
    styles.add(ParagraphStyle(
        name='Subtitulo',
        fontName='Helvetica-Bold',
        fontSize=11,
        textColor=colors.HexColor('#2c5282'),
        spaceBefore=5*mm,
        spaceAfter=3*mm,
    ))

    # Estilo para diagnóstico
    styles.add(ParagraphStyle(
        name='Diagnostico',
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.black,
        alignment=TA_LEFT,
        spaceBefore=1*mm,
        spaceAfter=1*mm,
        leftIndent=5*mm,
    ))

    return styles


def formatear_fecha(fecha):
    """Formatea la fecha para mostrar en el PDF."""
    if pd.isna(fecha):
        return ""
    if isinstance(fecha, str):
        return fecha
    try:
        return fecha.strftime("%d/%m/%Y")
    except:
        return str(fecha)


def formatear_rut(rut):
    """Asegura que el RUT tenga formato correcto."""
    if pd.isna(rut):
        return ""
    return str(rut).strip()


def obtener_logo_establecimiento(comuna):
    """Obtiene la ruta del logo según la comuna."""
    if pd.isna(comuna):
        return None
    comuna_upper = str(comuna).upper().strip()
    if comuna_upper in LOGOS_ESTABLECIMIENTO:
        logo_path = os.path.join(LOGOS_DIR, LOGOS_ESTABLECIMIENTO[comuna_upper])
        if os.path.exists(logo_path):
            return logo_path
    return None


def obtener_firma_oftalmologo(nombre_oftalmologo):
    """Obtiene la ruta de la firma del oftalmólogo."""
    if pd.isna(nombre_oftalmologo):
        return None
    nombre_upper = str(nombre_oftalmologo).upper().strip()
    for key, firma in FIRMAS_OFTALMOLOGO.items():
        if key in nombre_upper:
            firma_path = os.path.join(FIRMAS_DIR, firma)
            if os.path.exists(firma_path):
                return firma_path
    return None


def obtener_firma_tmo(nombre_tmo=None):
    """Obtiene la ruta de la firma del TMO."""
    # Por defecto usar firma de Felipe Rojas
    firma_default = os.path.join(FIRMAS_DIR, "firma_tmo_felipe_rojas.jpg")

    if nombre_tmo and not pd.isna(nombre_tmo):
        nombre_upper = str(nombre_tmo).upper().strip()
        for key, firma in FIRMAS_TMO.items():
            if key in nombre_upper:
                firma_path = os.path.join(FIRMAS_DIR, firma)
                if os.path.exists(firma_path):
                    return firma_path

    if os.path.exists(firma_default):
        return firma_default
    return None


def generar_pdf(paciente, output_path, styles):
    """Genera el PDF para un paciente."""

    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=1.5*cm,
        bottomMargin=2*cm,
    )

    elements = []

    # === OBTENER DATOS DEL PACIENTE PRIMERO (para logo establecimiento) ===
    nombre = str(paciente.get('NOMBRE PACIENTE', '')).strip() if not pd.isna(paciente.get('NOMBRE PACIENTE')) else ''
    rut = formatear_rut(paciente.get('RUT', ''))
    edad = str(int(paciente.get('EDAD', 0))) if not pd.isna(paciente.get('EDAD')) else ''
    fecha = formatear_fecha(paciente.get('FECHA', ''))
    institucion = str(paciente.get('ESTABLECIMIENTO', '')).strip() if not pd.isna(paciente.get('ESTABLECIMIENTO')) else ''
    comuna = str(paciente.get('COMUNA', '')).strip() if not pd.isna(paciente.get('COMUNA')) else ''

    # Logo del establecimiento
    logo_establecimiento = obtener_logo_establecimiento(comuna)

    # === ENCABEZADO ===
    logo_retidiag = os.path.join(LOGOS_DIR, "logo_retidiag.jpg")

    # Columna izquierda: datos de la empresa
    empresa_text = """
    <font size="8" color="#333333">www.retidiag.com<br/>
    Hernando de Aguirre 128 Of. 904<br/>
    Fono: 24816886/7<br/>
    Providencia, Santiago</font>
    """

    # Logo Retidiag
    if os.path.exists(logo_retidiag):
        logo_retidiag_img = Image(logo_retidiag, width=4*cm, height=1.2*cm)
    else:
        logo_retidiag_img = Paragraph("RETIDIAG", styles['Titulo'])

    # Logo del establecimiento en el encabezado
    if logo_establecimiento:
        logo_est_img = Image(logo_establecimiento, width=2.5*cm, height=1.8*cm)
        # Tabla con 3 columnas: empresa, logo retidiag, logo establecimiento
        encabezado_table = Table(
            [[Paragraph(empresa_text, styles['Empresa']), logo_retidiag_img, logo_est_img]],
            colWidths=[9*cm, 5*cm, 3*cm]
        )
        encabezado_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
    else:
        # Tabla con 2 columnas: empresa, logo retidiag
        encabezado_table = Table(
            [[Paragraph(empresa_text, styles['Empresa']), logo_retidiag_img]],
            colWidths=[12*cm, 5*cm]
        )
        encabezado_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

    elements.append(encabezado_table)
    elements.append(Spacer(1, 8*mm))

    # === TÍTULO ===
    elements.append(Paragraph("INFORME RETINOGRÁFICO", styles['Titulo']))

    # === DATOS DEL PACIENTE ===
    # Tabla de datos del paciente
    datos_paciente = [
        [Paragraph("<b>Nombre:</b>", styles['Etiqueta']),
         Paragraph(nombre, styles['Valor']),
         Paragraph("<b>Fecha Exámen:</b>", styles['Etiqueta']),
         Paragraph(fecha, styles['Valor'])],
        [Paragraph("<b>RUT:</b>", styles['Etiqueta']),
         Paragraph(rut, styles['Valor']),
         Paragraph("<b>Edad:</b>", styles['Etiqueta']),
         Paragraph(f"{edad} años" if edad else "", styles['Valor'])],
        [Paragraph("<b>Institución:</b>", styles['Etiqueta']),
         Paragraph(institucion, styles['Valor']),
         "", ""],
    ]

    datos_table = Table(datos_paciente, colWidths=[2.5*cm, 7*cm, 3*cm, 4*cm])
    datos_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))

    elements.append(datos_table)

    elements.append(Spacer(1, 8*mm))

    # === LÍNEA SEPARADORA ===
    elements.append(Table([['']], colWidths=[17*cm], rowHeights=[0.5*mm],
                         style=[('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#2c5282'))]))
    elements.append(Spacer(1, 5*mm))

    # === DIAGNÓSTICO ===
    resultado = str(paciente.get('RESULTADO FINAL', 'NORMAL')).upper().strip() if not pd.isna(paciente.get('RESULTADO FINAL')) else 'NORMAL'

    # Normalizar resultado
    if resultado in ['DG NORMAL', 'DGNORMAL']:
        resultado = 'DG NORMAL'
    elif 'NORMAL' in resultado and resultado != 'DG NORMAL':
        resultado = 'NORMAL'
    elif 'CATARATA' in resultado:
        resultado = 'CATARATA'
    elif resultado in ['RD', 'RETINOPATIA', 'RETINOPATÍA']:
        resultado = 'RD'
    elif resultado not in TEXTOS_DIAGNOSTICO:
        resultado = 'OTROS'

    # Texto introductorio
    elements.append(Paragraph(
        "Por medio de la evaluación realizada con cámara no midriática es posible informar que:",
        styles['Cuerpo']
    ))
    elements.append(Spacer(1, 3*mm))

    # Textos del diagnóstico
    textos = TEXTOS_DIAGNOSTICO.get(resultado, TEXTOS_DIAGNOSTICO['OTROS'])
    for texto in textos:
        if texto:
            elements.append(Paragraph(texto, styles['Diagnostico']))

    # Observaciones adicionales
    observaciones = paciente.get('OBSERVACIONES', '')
    if not pd.isna(observaciones) and str(observaciones).strip():
        elements.append(Spacer(1, 3*mm))
        elements.append(Paragraph(f"<b>Observaciones:</b> {observaciones}", styles['Diagnostico']))

    # Detalles OD/OI para todos los diagnósticos (siempre mostrar)
    detalle_od = paciente.get('DETALLE OD', '')
    detalle_oi = paciente.get('DETALLE OI', '')

    # Mostrar siempre los detalles de cada ojo
    detalle_od_texto = str(detalle_od).strip() if not pd.isna(detalle_od) and str(detalle_od).strip() else "Sin observaciones"
    detalle_oi_texto = str(detalle_oi).strip() if not pd.isna(detalle_oi) and str(detalle_oi).strip() else "Sin observaciones"

    elements.append(Spacer(1, 2*mm))
    elements.append(Paragraph(f"- Ojo Derecho (OD): {detalle_od_texto}", styles['Diagnostico']))
    elements.append(Paragraph(f"- Ojo Izquierdo (OI): {detalle_oi_texto}", styles['Diagnostico']))

    elements.append(Spacer(1, 5*mm))

    # === SUGERENCIAS ===
    elements.append(Paragraph("<b>SUGERENCIAS</b>", styles['Subtitulo']))

    sugerencias = SUGERENCIAS.get(resultado, SUGERENCIAS['OTROS'])
    for sugerencia in sugerencias:
        elements.append(Paragraph(sugerencia, styles['Diagnostico']))

    # Derivación
    derivacion = paciente.get('Derivacion', '')
    if not pd.isna(derivacion) and str(derivacion).strip():
        elements.append(Spacer(1, 2*mm))
        elements.append(Paragraph(f"<b>Derivación:</b> {derivacion}", styles['Diagnostico']))

    elements.append(Spacer(1, 5*mm))

    # === CIERRE ===
    elements.append(Paragraph("Es todo cuanto se puede informar.", styles['Cuerpo']))

    elements.append(Spacer(1, 10*mm))

    # === FIRMAS ===
    firma_tmo = obtener_firma_tmo()
    oftalmologo = paciente.get('OFTALMOLOGO', '')
    firma_oftalmologo = obtener_firma_oftalmologo(oftalmologo)

    # Determinar tipo de firma según resultado
    # DG NORMAL, RD, OTROS: solo firma de oftalmólogo
    # NORMAL, CATARATA: solo firma de TMO
    solo_oftalmologo = resultado in ['DG NORMAL', 'RD', 'OTROS']

    if solo_oftalmologo:
        # Solo firma de Oftalmólogo
        if firma_oftalmologo:
            firma_img = Image(firma_oftalmologo, width=4*cm, height=2.5*cm)
        else:
            firma_img = Spacer(1, 2.5*cm)

        firma_table = Table(
            [[firma_img],
             [Paragraph("Médico Oftalmólogo", styles['Valor'])]],
            colWidths=[4*cm]
        )
    else:
        # Solo firma de TMO (NORMAL, CATARATA)
        if firma_tmo:
            firma_img = Image(firma_tmo, width=4*cm, height=2.5*cm)
        else:
            firma_img = Spacer(1, 2.5*cm)

        firma_table = Table(
            [[firma_img],
             [Paragraph("Tecnólogo Médico", styles['Valor'])]],
            colWidths=[4*cm]
        )

    firma_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    # Centrar la tabla de firma en el documento
    firma_container = Table(
        [[firma_table]],
        colWidths=[17*cm]
    )
    firma_container.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))

    elements.append(firma_container)

    # Construir el PDF
    doc.build(elements)
    return True


def limpiar_nombre_archivo(nombre):
    """Limpia el nombre para usarlo como nombre de archivo."""
    if pd.isna(nombre):
        return "sin_nombre"
    # Reemplazar caracteres no válidos
    nombre = str(nombre).strip()
    caracteres_invalidos = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    for char in caracteres_invalidos:
        nombre = nombre.replace(char, '_')
    return nombre


def procesar_excel(excel_path, carpeta_salida=None):
    """Procesa el archivo Excel y genera los PDFs."""

    print(f"\n{'='*60}")
    print("GENERADOR DE INFORMES RETINOGRÁFICOS - RETIDIAG")
    print(f"{'='*60}\n")

    if not os.path.exists(excel_path):
        print(f"ERROR: No se encontró el archivo: {excel_path}")
        return False

    print(f"Leyendo archivo: {excel_path}")

    # Leer el Excel
    try:
        df = pd.read_excel(excel_path, sheet_name='INPUT', engine='openpyxl')
        # Limpiar espacios en nombres de columnas
        df.columns = df.columns.str.strip()
    except Exception as e:
        print(f"ERROR al leer el archivo: {e}")
        return False

    # Verificar columnas necesarias
    columnas_requeridas = ['NOMBRE PACIENTE', 'RUT', 'RESULTADO FINAL']
    for col in columnas_requeridas:
        if col not in df.columns:
            print(f"ERROR: Falta la columna '{col}' en el archivo Excel")
            return False

    # Filtrar filas válidas (que tengan nombre de paciente)
    df = df[df['NOMBRE PACIENTE'].notna()]

    print(f"Pacientes encontrados: {len(df)}")

    # Obtener comuna, establecimiento y fecha para nombre de carpeta
    comuna = ""
    establecimiento = ""
    fecha_examen = ""

    if len(df) > 0:
        # Obtener la comuna (usar la primera o la más común)
        if 'COMUNA' in df.columns:
            comuna = df['COMUNA'].mode().iloc[0] if not df['COMUNA'].mode().empty else df['COMUNA'].iloc[0]
            if pd.isna(comuna):
                comuna = "Sin_Comuna"
            else:
                comuna = str(comuna).strip()

        # Obtener el establecimiento (usar el primero o el más común)
        if 'ESTABLECIMIENTO' in df.columns:
            establecimiento = df['ESTABLECIMIENTO'].mode().iloc[0] if not df['ESTABLECIMIENTO'].mode().empty else df['ESTABLECIMIENTO'].iloc[0]
            if pd.isna(establecimiento):
                establecimiento = "Sin_Establecimiento"
            else:
                establecimiento = str(establecimiento).strip()

        # Obtener la fecha del examen (usar la primera o la más común)
        if 'FECHA' in df.columns:
            fecha_valor = df['FECHA'].iloc[0]
            if not pd.isna(fecha_valor):
                if isinstance(fecha_valor, str):
                    fecha_examen = fecha_valor.replace('/', '-').replace(' ', '_')[:10]
                else:
                    try:
                        fecha_examen = fecha_valor.strftime("%Y-%m-%d")
                    except:
                        fecha_examen = str(fecha_valor)[:10]
            else:
                fecha_examen = datetime.now().strftime("%Y-%m-%d")

    # Limpiar nombres para usar como carpetas
    comuna_limpia = limpiar_nombre_archivo(comuna)
    establecimiento_limpio = limpiar_nombre_archivo(establecimiento)

    # Crear carpeta con establecimiento y fecha
    nombre_carpeta_establecimiento = f"{establecimiento_limpio}_{fecha_examen}"

    # Crear estructura: informes retidiag / COMUNA / Establecimiento_Fecha
    if carpeta_salida:
        output_dir = os.path.join(carpeta_salida, comuna_limpia, nombre_carpeta_establecimiento)
    else:
        output_dir = os.path.join(OUTPUT_DIR, comuna_limpia, nombre_carpeta_establecimiento)

    os.makedirs(output_dir, exist_ok=True)
    print(f"Comuna: {comuna}")
    print(f"Establecimiento: {establecimiento}")
    print(f"Fecha examen: {fecha_examen}")
    print(f"Carpeta de salida: {output_dir}\n")

    # Crear estilos
    styles = crear_estilos()

    # Contadores
    exitosos = 0
    errores = 0

    # Procesar cada paciente
    for idx, paciente in df.iterrows():
        nombre = paciente.get('NOMBRE PACIENTE', f'paciente_{idx}')
        nombre_archivo = limpiar_nombre_archivo(nombre)
        resultado = str(paciente.get('RESULTADO FINAL', 'NORMAL')).strip() if not pd.isna(paciente.get('RESULTADO FINAL')) else 'NORMAL'

        # Crear subcarpeta por resultado
        resultado_dir = os.path.join(output_dir, resultado.replace(' ', '_'))
        os.makedirs(resultado_dir, exist_ok=True)

        # Nombre del archivo PDF
        pdf_filename = f"{nombre_archivo}.pdf"
        pdf_path = os.path.join(resultado_dir, pdf_filename)

        try:
            generar_pdf(paciente.to_dict(), pdf_path, styles)
            print(f"✓ {nombre} -> {resultado}")
            exitosos += 1
        except Exception as e:
            print(f"✗ ERROR con {nombre}: {e}")
            errores += 1

    print(f"\n{'='*60}")
    print(f"RESUMEN:")
    print(f"  - PDFs generados exitosamente: {exitosos}")
    print(f"  - Errores: {errores}")
    print(f"  - Ubicación: {output_dir}")
    print(f"{'='*60}\n")

    # Generar archivo resumen de pacientes
    generar_resumen_pacientes(df, establecimiento, fecha_examen, output_dir)

    return True


def generar_resumen_pacientes(df, establecimiento, fecha_examen, output_dir):
    """Genera el archivo Excel resumen de pacientes con formato y logo."""

    print("Generando resumen de pacientes...")

    # Crear nuevo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Pacientes"

    # Colores
    azul_header = PatternFill(start_color="2c5282", end_color="2c5282", fill_type="solid")
    azul_claro = PatternFill(start_color="e8f0fe", end_color="e8f0fe", fill_type="solid")
    blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Estilos
    title_font = Font(bold=True, size=16, color="2c5282")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin', color='2c5282'),
        right=Side(style='thin', color='2c5282'),
        top=Side(style='thin', color='2c5282'),
        bottom=Side(style='thin', color='2c5282')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Agregar logo de Retidiag (filas 1-4)
    logo_path = os.path.join(LOGOS_DIR, "logo_retidiag.jpg")
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 150
        img.height = 45
        ws.add_image(img, 'A1')

    # Ajustar altura de filas para el logo
    for row in range(1, 5):
        ws.row_dimensions[row].height = 15

    # Fila 6: Nombre del establecimiento
    ws.merge_cells('A6:J6')
    cell_estab = ws.cell(row=6, column=1, value=establecimiento)
    cell_estab.font = title_font
    cell_estab.alignment = center_align
    ws.row_dimensions[6].height = 25

    # Fila 7: Encabezados
    encabezados = ["N°", "Fecha", "Centro", "Rut", "Nombre", "Edad", "Diagnóstico", "Ojo Derecho", "Ojo Izquierdo", "Observación"]
    for col, encabezado in enumerate(encabezados, 1):
        cell = ws.cell(row=7, column=col, value=encabezado)
        cell.font = header_font
        cell.fill = azul_header
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[7].height = 20

    # Fila 8 en adelante: Datos de pacientes
    for idx, (_, paciente) in enumerate(df.iterrows(), 1):
        row = idx + 7  # Empezar en fila 8

        # Alternar colores de fila
        fill = azul_claro if idx % 2 == 0 else blanco

        # N°
        cell = ws.cell(row=row, column=1, value=idx)
        cell.border = thin_border
        cell.fill = fill
        cell.font = data_font
        cell.alignment = center_align

        # Fecha
        fecha = paciente.get('FECHA', '')
        if not pd.isna(fecha):
            if hasattr(fecha, 'strftime'):
                fecha = fecha.strftime("%d/%m/%Y")
            else:
                fecha = str(fecha)[:10]
        else:
            fecha = ""
        cell = ws.cell(row=row, column=2, value=fecha)
        cell.border = thin_border
        cell.fill = fill
        cell.font = data_font
        cell.alignment = center_align

        # Centro
        centro = str(paciente.get('ESTABLECIMIENTO', '')).strip() if not pd.isna(paciente.get('ESTABLECIMIENTO')) else ''
        cell = ws.cell(row=row, column=3, value=centro)
        cell.border = thin_border
        cell.fill = fill
        cell.font = data_font
        cell.alignment = left_align

        # Rut
        rut = str(paciente.get('RUT', '')).strip() if not pd.isna(paciente.get('RUT')) else ''
        cell = ws.cell(row=row, column=4, value=rut)
        cell.border = thin_border
        cell.fill = fill
        cell.font = data_font
        cell.alignment = center_align

        # Nombre
        nombre = str(paciente.get('NOMBRE PACIENTE', '')).strip() if not pd.isna(paciente.get('NOMBRE PACIENTE')) else ''
        cell = ws.cell(row=row, column=5, value=nombre)
        cell.border = thin_border
        cell.fill = fill
        cell.font = data_font
        cell.alignment = left_align

        # Edad
        edad = int(paciente.get('EDAD', 0)) if not pd.isna(paciente.get('EDAD')) else ''
        cell = ws.cell(row=row, column=6, value=edad)
        cell.border = thin_border
        cell.fill = fill
        cell.font = data_font
        cell.alignment = center_align

        # Diagnóstico (de RESULTADO FINAL)
        diagnostico = str(paciente.get('RESULTADO FINAL', '')).strip() if not pd.isna(paciente.get('RESULTADO FINAL')) else ''
        cell = ws.cell(row=row, column=7, value=diagnostico)
        cell.border = thin_border
        cell.fill = fill
        cell.font = data_font
        cell.alignment = center_align

        # Ojo Derecho
        ojo_derecho = str(paciente.get('DETALLE OD', '')).strip() if not pd.isna(paciente.get('DETALLE OD')) else ''
        cell = ws.cell(row=row, column=8, value=ojo_derecho)
        cell.border = thin_border
        cell.fill = fill
        cell.font = data_font
        cell.alignment = center_align

        # Ojo Izquierdo
        ojo_izquierdo = str(paciente.get('DETALLE OI', '')).strip() if not pd.isna(paciente.get('DETALLE OI')) else ''
        cell = ws.cell(row=row, column=9, value=ojo_izquierdo)
        cell.border = thin_border
        cell.fill = fill
        cell.font = data_font
        cell.alignment = center_align

        # Observación (desde campo Derivacion)
        derivacion = str(paciente.get('Derivacion', '')).strip() if not pd.isna(paciente.get('Derivacion')) else ''
        cell = ws.cell(row=row, column=10, value=derivacion)
        cell.border = thin_border
        cell.fill = fill
        cell.font = data_font
        cell.alignment = left_align

    # Ajustar ancho de columnas
    anchos = {'A': 5, 'B': 12, 'C': 30, 'D': 12, 'E': 28, 'F': 6, 'G': 12, 'H': 12, 'I': 12, 'J': 25}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    # Guardar archivo
    establecimiento_limpio = limpiar_nombre_archivo(establecimiento)
    nombre_archivo = f"{establecimiento_limpio}_{fecha_examen}.xlsx"
    ruta_archivo = os.path.join(output_dir, nombre_archivo)

    wb.save(ruta_archivo)
    print(f"✓ Resumen guardado: {nombre_archivo}")


def main():
    """Función principal."""
    if len(sys.argv) < 2:
        # Usar archivo por defecto
        excel_path = "/Users/magda/Downloads/Plantilla para crear informes PDF de FO 2026.xlsm"
        if not os.path.exists(excel_path):
            print("Uso: python generar_informes.py <archivo_excel.xlsx> [carpeta_salida]")
            print("\nEjemplo:")
            print("  python generar_informes.py datos_pacientes.xlsx")
            print("  python generar_informes.py datos_pacientes.xlsx ./mis_informes")
            sys.exit(1)
    else:
        excel_path = sys.argv[1]

    carpeta_salida = sys.argv[2] if len(sys.argv) > 2 else None

    procesar_excel(excel_path, carpeta_salida)


if __name__ == "__main__":
    main()
